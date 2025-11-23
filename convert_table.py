#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import os
from typing import Union, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _process_single_md_to_sheet(ws, content: str, sheet_name: str = '资产负债表'):
    """
    处理单个MD文件内容并写入到指定的worksheet

    Args:
        ws: openpyxl worksheet对象
        content: MD文件内容
        sheet_name: sheet名称

    Returns:
        处理成功返回True,否则返回False
    """
    # 提取HTML表格部分
    table_match = re.search(r'<table.*?</table>', content, re.DOTALL)
    if not table_match:
        return False

    ws.title = sheet_name[:31]  # Excel sheet名称最多31字符
    html_table = table_match.group(0)
    table_start = table_match.start()
    table_end = table_match.end()

    # 提取表格前的内容(去除空行)
    before_table = content[:table_start].strip()
    before_lines = [line.strip() for line in before_table.split('\n') if line.strip()]

    # 提取表格后的内容(去除空行)
    after_table = content[table_end:].strip()
    after_lines = [line.strip() for line in after_table.split('\n') if line.strip()]

    row_idx = 1

    # 1. 输出表格前的内容
    for line in before_lines:
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row_idx += 1

    # 2. 解析并输出HTML表格
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html_table, re.DOTALL)

    # 存储合并单元格的信息
    merged_cells = []
    table_start_row = row_idx  # 记录表格起始行

    # 跟踪每个位置的占用情况 {(row, col): True}
    occupied = {}

    for row_html in rows:
        # 提取所有单元格
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL)
        # 重新查找完整的td标签以获取属性
        td_tags = re.findall(r'<td[^>]*>.*?</td>', row_html, re.DOTALL)

        col_idx = 1
        for i, cell_content in enumerate(cells):
            td_tag = td_tags[i]

            # 找到第一个未被占用的列
            while occupied.get((row_idx, col_idx), False):
                col_idx += 1

            # 提取rowspan和colspan
            rowspan_match = re.search(r'rowspan="(\d+)"', td_tag)
            colspan_match = re.search(r'colspan="(\d+)"', td_tag)

            rowspan = int(rowspan_match.group(1)) if rowspan_match else 1
            colspan = int(colspan_match.group(1)) if colspan_match else 1

            # 清理HTML标签
            cell_text = re.sub(r'<[^>]+>', '', cell_content).strip()

            # 写入单元格
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border

            # 为表格的第一行和第二行添加背景色（表头）
            if row_idx <= table_start_row + 1:
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # 标记当前单元格占用的所有位置
            for r_offset in range(rowspan):
                for c_offset in range(colspan):
                    occupied[(row_idx + r_offset, col_idx + c_offset)] = True

            # 记录合并信息
            if rowspan > 1 or colspan > 1:
                start_col = get_column_letter(col_idx)
                end_col = get_column_letter(col_idx + colspan - 1)
                end_row = row_idx + rowspan - 1
                merged_cells.append(f'{start_col}{row_idx}:{end_col}{end_row}')

            col_idx += 1

        row_idx += 1

    # 应用合并单元格
    for merged_range in merged_cells:
        try:
            ws.merge_cells(merged_range)
        except Exception as e:
            print(f"合并失败 {merged_range}: {e}")

    # 3. 输出表格后的内容
    for line in after_lines:
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row_idx += 1

    # 调整列宽
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    # 调整行高
    for row in ws.rows:
        ws.row_dimensions[row[0].row].height = 30

    return True


def convert_md_to_excel(
    md_files: Union[str, List[str]],
    output_path: Optional[str] = None
) -> List[str]:
    """
    将一个或多个MD文件转换为Excel文件

    Args:
        md_files: MD文件路径(字符串)或MD文件路径列表
        output_path: 输出Excel文件路径,如果为None则自动生成

    Returns:
        生成的Excel文件路径列表

    Raises:
        FileNotFoundError: 当MD文件不存在时
        ValueError: 当MD文件列表为空时
    """
    # 统一转换为列表
    if isinstance(md_files, str):
        md_files = [md_files]

    if not md_files:
        raise ValueError("MD文件列表不能为空")

    # 验证所有文件都存在
    for md_file in md_files:
        if not os.path.exists(md_file):
            raise FileNotFoundError(f"MD文件不存在: {md_file}")

    # 创建Excel工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    processed_count = 0

    # 处理每个MD文件
    for md_file in md_files:
        # 读取MD文件内容
        try:
            with open(md_file, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            print(f"读取文件失败 {md_file}: {e}")
            continue

        # 生成sheet名称(使用文件名,去除扩展名)
        sheet_name = os.path.splitext(os.path.basename(md_file))[0]

        # 创建新的worksheet
        ws = wb.create_sheet(title=sheet_name[:31])

        # 处理MD内容到sheet
        success = _process_single_md_to_sheet(ws, content, sheet_name)

        if success:
            processed_count += 1
            print(f"成功处理: {md_file} -> Sheet: {ws.title}")
        else:
            print(f"未找到表格: {md_file}")
            # 如果处理失败,删除这个sheet
            wb.remove(ws)

    # 如果没有成功处理任何文件,返回空列表
    if processed_count == 0:
        print("警告: 没有成功处理任何MD文件")
        return []

    # 确定输出路径
    if output_path is None:
        # 使用第一个MD文件的目录和合并后的文件名
        input_dir = os.path.dirname(md_files[0])
        output_path = os.path.join(input_dir, "merged_output.xlsx")

    # 保存Excel文件
    wb.save(output_path)
    print(f"\n成功转换为Excel: {output_path}")
    print(f"处理文件数: {processed_count}/{len(md_files)}")
    print(f"Sheet数量: {len(wb.sheetnames)}")

    return [output_path]


# 兼容原有的直接运行方式
if __name__ == "__main__":
    # 文件路径
    file_path = r'E:\OneDrive\project_code\output\page_005.md'  # 大普微测试
    file_path = r"e:\OneDrive\project_code\output\tmpgkh6u3z0.md"

    # 使用新函数
    result = convert_md_to_excel(file_path)
    if result:
        print(f"\n输出文件: {result[0]}")
